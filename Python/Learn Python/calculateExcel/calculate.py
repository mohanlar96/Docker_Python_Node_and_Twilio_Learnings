import openpyxl as xl
wb=xl.load_workbook('transactions.xlsx')
sheet=wb['Sheet1']
cell=sheet['a1']
cell=sheet.cell(1,1)
print(cell.value)

for row in range(1,sheet.max_row+1):
    cell=sheet.cell(row,3)
    corrected_price=float(cell.value)*0.9
    corrected_price_cell=sheet.cell(row,4)
    corrected_price_cell.value=corrected_price

wb.save('transactions2.xlsx')